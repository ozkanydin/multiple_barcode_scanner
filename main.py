from pyzbar.pyzbar import decode
from PIL import Image
from openpyxl import Workbook,load_workbook
import os,cv2

i = 0





wb = Workbook()

ws = wb.active
ws.title = "İlk Çalışma Alanı"


print(wb.sheetnames)
ws.append([" SERI NO ","FOTO YOLLARI"])
dosyalar = os.listdir("..\\fotolar")
sn = []
saydir = 0
for dosyax in dosyalar:
    konum = "..\\fotolar\\" + dosyax
    konumm = "..\\fotolar\\"
    fd = konumm + str(saydir) + ".png"
    os.rename(konumm + dosyax,fd )

    saydir = saydir + 1
dosyalari = os.listdir("..\\fotolar")
for dosya in dosyalari:

    konum = "..\\fotolar\\"+dosya
    konumm = "..\\fotolar\\"
    new_konum = konum
    fotocv2 = cv2.imread(new_konum)

    i = 0
    for bar in decode(Image.open(new_konum)):

        if len(bar.data) == 14 and bar.type == "CODE128":
            i = i + 1
            ws.append([bar.data, new_konum])
            sn.append(bar.data)
            (x, y, w, h) = bar.rect
            #fotocv2 = cv2.line(fotocv2, (x,y) , (x+w,y+h)
            #bar.data
            # fotocv2 = cv2.rectangle(fotocv2, (x, y), (x + w, y + h), (255, 0, 0), 15)
            fotocv2 =  cv2.putText(fotocv2, "OKUNDU", (x+w,y+h),
                            cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 5)


    if i < 20:
        fotocv2 = cv2.putText(fotocv2, str(i) + " adet okundu", (150, 150),
                              cv2.FONT_HERSHEY_SIMPLEX, 1.5, (0, 255, 0), 3)
        cv2.imwrite("..\\okunmayan\\"+dosya+".png",fotocv2)




        # if len(bar.data) == 14 and bar.type == "CODE128" :
        #     # (x, y, w, h) = bar.rect
        #     # cv2.rectangle(image, (x, y), (x + w, y + h), (255, 0, 0), 5)
        #
        #
        #
        #     #print("POLY:",bar.rect[1])
        #     #print("BAR : ",bar)
        #     ws.append([bar.data,konum ])
        #     sn.append(bar.data)

    # cv2.imshow("Image", image)
    #
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()

    wb.save("..\\dosya.xlsx")

for x in sn:
    print(x)
print(len(sn)," adet kayıt eklendi.")

input("Bitti...")
